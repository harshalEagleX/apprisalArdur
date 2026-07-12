"""
tools/xlsx_to_checklist.py — regenerate an AMC checklist in the NATIVE language
format (final_shalqccore.md / Annex B §3 Step 0) straight from the source Excel.

"Our new way": the compiled item needs only {item_id, section, check_text,
reject_text}. No rule_id, no hand-authored `sources`, no per-check code — the
AMC's own words ARE the rule. This converter reads the richest sheet in the
Equity Solutions workbook (`Appraisal QC Checklist -Detailed.xlsx` → `Sheet1`,
which carries explicit "Reject as:" wording in the trailing columns) and writes
`readme/exampleAMC/checklist_equitysolutions.yaml`.

Usage:  PYTHONPATH=. python tools/xlsx_to_checklist.py
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import yaml

_DIR = Path(__file__).parent.parent / "readme" / "exampleAMC"
_SRC = _DIR / "Appraisal QC Checklist -Detailed.xlsx"
_OUT = _DIR / "checklist_equitysolutions.yaml"
_LEGACY_OUT = _DIR / "qc_rejection_catalog.yaml"

# section → legacy rule_id prefix (matches app/rules/catalog.py _section_of).
_PREFIX = {
    "order": "ORD", "subject": "S", "contract": "C", "neighborhood": "N",
    "site": "ST", "improvements": "I", "sales_comparison": "SCA",
    "prior_sales": "PSH", "reconciliation": "R", "cost": "CA", "income": "INC",
    "pud": "PUD", "signature": "SIG", "mc_1004": "MCA", "uspap": "USPAP",
    "exhibits": "EXH",
}
_NARR_RX = re.compile(r"(description|commentary|narrative|analysis|summary|boundaries|remarks)", re.I)

# Section-header text (col B, no id in col A) → our canonical section name.
_SECTION_MAP = [
    ("SUBJECT", "subject"), ("CONTRACT", "contract"), ("NEIGHBORHOOD", "neighborhood"),
    ("SITE", "site"), ("IMPROVEMENT", "improvements"), ("SALES COMPARISON", "sales_comparison"),
    ("PRIOR SALES", "prior_sales"), ("RECONCILIATION", "reconciliation"),
    ("COST APPROACH", "cost"), ("INCOME APPROACH", "income"), ("PUD", "pud"),
    ("SIGNATURE", "signature"), ("MARKET CONDITIONS", "mc_1004"), ("USPAP", "uspap"),
    ("PHOTOGRAPH", "exhibits"), ("FLOORPLAN", "exhibits"), ("SKETCH", "exhibits"),
    ("MAPS", "exhibits"), ("ADDITIONAL PAGES", "exhibits"),
    ("POINTERS", "order"),
]

_REJECT_RX = re.compile(r"^\s*reject\s*as\s*:?\s*", re.I)
_ID_RX = re.compile(r"[0-9]+|[A-G]\b")


def _s(x) -> str:
    return "" if x is None else str(x).replace("\n", " ").replace("  ", " ").strip()


def _is_section_header(a: str, b: str, c: str, d: str) -> Optional[str]:
    if a or not b:
        return None
    up = b.upper()
    # a header row has a section-ish title and no requirement text of its own.
    for key, sec in _SECTION_MAP:
        if key in up and (not c or up.endswith(("SECTION", "APPROACH", "PAGE",
                          "ADDENDUM", "INFORMATION", "SKETCH", "MAPS", "REQUIRED",
                          "VALUE", "PROJECTS"))):
            return sec
    return None


def _clean_id(a: str) -> str:
    m = _ID_RX.search(a)
    return m.group(0) if m else re.sub(r"[^A-Za-z0-9]", "", a)[:6]


def _extract_reject(cells: List[str]) -> Optional[str]:
    """The verbatim AMC reject wording from any trailing cell that opens with
    'Reject as:'. Multiple → the longest (most complete) one."""
    found = []
    for c in cells:
        if _REJECT_RX.match(c):
            w = _REJECT_RX.sub("", c).strip(' "“”')
            if len(w) > 4:
                found.append(w)
    return max(found, key=len) if found else None


def _conditions(cells: List[str]) -> List[str]:
    """The 'If …' trigger clauses (col D/F/G/H) — context the judge evaluates,
    kept out of reject_text."""
    out = []
    for c in cells:
        if _REJECT_RX.match(c):
            continue
        if re.match(r"^\s*(if|when|where|is\b|are\b)", c, re.I) and len(c) > 6:
            out.append(c)
    return out


def parse(src: Path = _SRC) -> List[Dict]:
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Sheet1"]
    items: List[Dict] = []
    section = "order"
    current: Optional[Dict] = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [(_s(v)) for v in (list(row) + [None] * 9)[:9]]
        a, b, c = vals[0], vals[1], vals[2]
        trailing = vals[3:]                     # D..I: conditions + reject wording
        if not any(vals):
            continue

        sec = _is_section_header(a, b, c, trailing[0] if trailing else "")
        if sec:
            section = sec
            current = None
            continue

        if a:                                   # a new checklist item
            reject = _extract_reject(trailing)
            conds = _conditions(trailing)
            parts = [p for p in (b, c) if p]
            check_text = " — ".join(parts)
            if conds:
                check_text += " | Triggers: " + "; ".join(conds)
            current = {
                "item_id": f"EQ-{_clean_id(a)}",
                "section": section,
                "item": b or c[:40],
                "check_text": check_text.strip(" —"),
                "reject_text": reject,
            }
            items.append(current)
        elif current is not None:               # continuation of the prior item
            if c:
                current["check_text"] += f" {c}"
            reject = _extract_reject(trailing)
            if reject and not current.get("reject_text"):
                current["reject_text"] = reject
            for cond in _conditions(trailing):
                current["check_text"] += f" | Trigger: {cond}"

    # de-dupe item_ids (some rows share a number, e.g. "15,"): suffix -b, -c…
    seen: Dict[str, int] = {}
    for it in items:
        base = it["item_id"]
        n = seen.get(base, 0)
        if n:
            it["item_id"] = f"{base}-{chr(ord('a') + n)}"
        seen[base] = n + 1
    return items


def _infer_check_type(section: str, item: str, check_text: str) -> str:
    if section == "exhibits":
        return "cross_modal"
    if _NARR_RX.search(item) or _NARR_RX.search(check_text):
        return "narrative"
    if section in ("order",) or re.search(r"match|engagement|order form", check_text, re.I):
        return "cross_document"
    return "presence"


def _legacy_sources(check_text: str) -> List[Dict]:
    """Best-effort appraisal source fields (heuristic label bind) so a restored
    legacy presence check gates on a real field instead of nothing."""
    from app.language.compiler import _heuristic_labels
    labels = _heuristic_labels(check_text, limit=2)
    return [{"doc": "appraisal", "fields": labels}] if labels else []


def to_legacy(items: List[Dict]) -> List[Dict]:
    """Re-emit the native items in the legacy rejection-catalog schema so the
    superseded rule engine + its tests keep working. rule_id is section-prefixed
    and sequential (N-1, SCA-2, I-1, …); hand-coded ids (S-1, R-1) are naturally
    skipped by the catalog interpreter's dedup."""
    seq: Dict[str, int] = {}
    out: List[Dict] = []
    for it in items:
        sec = it["section"]
        prefix = _PREFIX.get(sec, "CAT")
        seq[prefix] = seq.get(prefix, 0) + 1
        rid = f"{prefix}-{seq[prefix]}"
        ct = _infer_check_type(sec, it.get("item", ""), it["check_text"])
        out.append({
            "id": it["item_id"].replace("EQ-", ""),
            "rule_id": rid,
            "item": it.get("item", ""),
            "check_type": ct,
            "requirement": it["check_text"],
            "reject_as": [it["reject_text"]] if it.get("reject_text") else [],
            "sources": _legacy_sources(it["check_text"]),
        })
    return out


def main() -> None:
    items = parse()
    doc = {
        "meta": {
            "amc": "EQUITYSOLUTIONS",
            "source": _SRC.name + " → Sheet1",
            "format": "native (final_shalqccore.md Annex B §3 Step 0)",
            "count": len(items),
        },
        "items": items,
    }
    _OUT.write_text(yaml.safe_dump(doc, sort_keys=False, allow_unicode=True), encoding="utf-8")
    withrej = sum(1 for it in items if it.get("reject_text"))
    print(f"wrote {_OUT} — {len(items)} items ({withrej} with verbatim reject wording)")
    from collections import Counter
    print("sections:", dict(Counter(it["section"] for it in items)))

    # also (re)generate the legacy rejection catalog the old engine/tests expect.
    legacy = to_legacy(items)
    legacy_doc = {
        "meta": {"source": _SRC.name, "regenerated_from": "native checklist",
                 "note": "legacy schema for app/rules/catalog.py; superseded by the "
                         "language path (checklist_equitysolutions.yaml)",
                 "total_items": len(legacy)},
        "items": legacy,
    }
    _LEGACY_OUT.write_text(yaml.safe_dump(legacy_doc, sort_keys=False, allow_unicode=True), encoding="utf-8")
    print(f"wrote {_LEGACY_OUT} — {len(legacy)} legacy items "
          f"(rule_ids incl. {', '.join(sorted({l['rule_id'] for l in legacy})[:6])}…)")


if __name__ == "__main__":
    main()
